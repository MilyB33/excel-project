from customtkinter import CTk, CTkFrame, CTkEntry, CTkButton, CTkCheckBox, filedialog, END
import os
import constant
from windows import Window
from tkinter import messagebox, BooleanVar, simpledialog
from openpyxl import Workbook
from openpyxl.styles import PatternFill

class CSV_WINDOW(Window):
    __file_path__: str | None = None
    __file_name__: str | None = None
    __workbook__: Workbook | None = None

    def __init__(self, root: CTk):
        super().__init__("Import plików CSV", constant.MIN_WINDOW_WIDTH, constant.MIN_WINDOW_HEIGHT, root)

    def __create_layout__(self):
        frame = CTkFrame(self.window, bg_color='lightblue')
        frame.pack(expand=True)

        # Przycisk do powrotu
        go_back_button =CTkButton(frame ,text="<-- Powrót", command= self.__go_back__)
        go_back_button.grid(row=0, column=0, padx=10, pady=10, sticky="w")

        text_entry = CTkEntry(frame, width=300, placeholder_text="Wybierz plik")
        text_entry.grid(row=1, column=0, padx=10,pady=10)

        self.__is_first_row_head__: BooleanVar = BooleanVar()
        self.__is_align_content__: BooleanVar = BooleanVar()
        self.__is_sort__: BooleanVar = BooleanVar()


        is_first_row_head_checkbox = CTkCheckBox(frame, text="Pierwszy wiersz zawiera nazwy kolumn", variable=self.__is_first_row_head__, onvalue=True, offvalue=False)
        is_first_row_head_checkbox.grid(row=2, column=0, padx=10, pady=10, sticky="w")

        is_align_content_checkbox = CTkCheckBox(frame, text="Wyrównaj", variable=self.__is_align_content__, onvalue=True, offvalue=False)
        is_align_content_checkbox.grid(row=3, column=0, padx=10, pady=10, sticky="w")

        is_sort_checkbox = CTkCheckBox(frame, text="Sortuj", variable=self.__is_sort__, onvalue=True, offvalue=False)
        is_sort_checkbox.grid(row=4, column=0, padx=10, pady=10, sticky="w")

        choose_file_button = CTkButton(frame, text="Wybierz plik", command=lambda: self.__choose_file__(text_entry))
        choose_file_button.grid(row=5, column=0, padx=10, pady=10)

        import_data_button = CTkButton(frame, text="Zapisz", command= self.__import_data__)
        import_data_button.grid(row=6, column=0, padx=10, pady=10)

    def open_window(self):
        super().open_window()
        self.__create_layout__()

    def __choose_file__(self, text_entry: CTkEntry):
        try:
            file  = filedialog.askopenfile(mode='r', initialdir="./", filetypes=[("txt", '*.txt'),('CSV', '*.csv')])

            if file:
                file_path = os.path.abspath(file.name)
                self.__file_path__ = file_path
                self.__file_name__ = os.path.basename(file.name)
                text_entry.delete(0, END)
                text_entry.insert(0, str(file_path))
        except:
            messagebox.showerror(title="Błąd przy wyborze pliku", message="Błąd. Upewnij się że wszystkie okna excela są zamknięte")

    def __import_data__(self):
        if not self.__file_path__:
            messagebox.showerror(title="Błąd", message="Nie wybrano pliku")
            return

        workbook = Workbook()
        self.__workbook__ = workbook

        ws = workbook.active

        if ws and self.__file_name__:
            ws.title = f'{self.__file_name__}-worksheet-{len(workbook.worksheets) + 1}'
            self.__load_all_data__()

        is_first_row_head = self.__is_first_row_head__.get()
        is_align_content = self.__is_align_content__.get()
        is_sort = self.__is_sort__.get()

        if is_first_row_head:
            self.__color_head_row__()

        if is_align_content:
            self.__align_content__()

        if is_sort:
            self.__sort__()

        self.__on_safe__()

    def __load_all_data__(self):
        ws = self.__workbook__.active
        
        if self.__file_path__ and self.__workbook__ and ws:
            with open(self.__file_path__, 'r') as file:
                lines = file.readlines()

           
        for row_num, line in enumerate(lines, start=1):
            columns = line.strip().split('\t')  # Zakładamy, że dane są rozdzielone tabulatorami ('\t')
            for col_num, value in enumerate(columns, start=1):
               ws.cell(row=row_num, column=col_num, value=value)
    
    def __on_safe__(self):
        try: 
            if self.__workbook__:
                filename = simpledialog.askstring("Filename", "Podaj nazwę pliku do jakiego zapisać")

                if filename:
                    self.__workbook__.save(f'./results/{filename}.xlsx')
                    messagebox.showinfo(title="Zapisano", message="Poprawnie zapisano")
                else:
                    messagebox.showerror(title="Błąd zapisu", message="Nie wybrano nazwy pliku")
            else:
                raise Exception("No active workbook")
        except: 
            messagebox.showerror(title="Błąd zapisu", message="Brak aktywnego workbooka lub excel jest już otwarty. Upewnij się że wszystkie okna są zamknięte")

    def __color_head_row__(self):
        ws = self.__workbook__.active

        if ws:
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=1, column=col).value != '':
                    ws.cell(row=1, column=col).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    
    def __align_content__(self):
        ws = self.__workbook__.active

        if ws:
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column].width = adjusted_width

    def __sort__(self):
        ws = self.__workbook__.active

        if ws:
            ws.sort_column = 1