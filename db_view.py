from customtkinter import CTk, CTkFrame, CTkEntry, CTkButton,CTkLabel, CTkCheckBox, filedialog, END
from windows import Window
import constant
import mysql.connector
from tkinter import messagebox, StringVar, BooleanVar, simpledialog
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side
import os
from pathlib import Path

class DB_WINDOW(Window):
    __workbook__: Workbook | None = None
    __workbook_path__: str | None = None

    def __init__(self, root: CTk):
        super().__init__("Łączenie z bazą", constant.MIN_WINDOW_WIDTH, constant.MIN_WINDOW_HEIGHT, root)

    def __create_layout__(self):
        frame = CTkFrame(self.window, fg_color="transparent")
        frame.pack(expand=True)
        
        # Przycisk do powrotu
        go_back_button =CTkButton(frame ,text="<-- Powrót", command= self.__go_back__)
        go_back_button.grid(row=0, column=0, padx=10, pady=10, sticky="w")

        self.__create_connect_form__(frame)
        self.__create_style_form__(frame)

    def __create_connect_form__(self, rootFrame: CTkFrame):
        frame = CTkFrame(rootFrame)
        frame.grid(row=1, column=0)

        self.__user_name__ = StringVar(self.window, None)
        self.__user_password__ = StringVar(self.window, None)
        self.__host__ = StringVar(self.window, None)
        self.__database_name__ = StringVar(self.window, None)

        # username
        username_frame = CTkFrame(frame)
        username_frame.grid(row=0, column=0)

        user_name_label = CTkLabel(username_frame, text="Nazwa użytkownika:")
        user_name_label.grid(row=0, column=0)

        user_name_entry = CTkEntry(username_frame, width=200, textvariable=self.__user_name__, placeholder_text="Podaj nazwe użytkownika",)
        user_name_entry.grid(row=1, column=0, padx=10, pady=10)

        # password
        password_frame = CTkFrame(frame)
        password_frame.grid(row=0, column=1)

        password_label = CTkLabel(password_frame, text="Hasło użytkownika:")
        password_label.grid(row=0, column=0)

        user_password_entry = CTkEntry(password_frame, width=200, textvariable=self.__user_password__)
        user_password_entry.grid(row=1, column=0, padx=10, pady=10)

        # host
        host_frame = CTkFrame(frame)
        host_frame.grid(row=1, column=0, columnspan=2)

        host_label = CTkLabel(host_frame, text="Host:")
        host_label.grid(row=0, column=0)

        host_entry = CTkEntry(host_frame, width=400, textvariable=self.__host__)
        host_entry.grid(row=1, column=0, padx=10, pady=10)

        # database name
        database_name_frame = CTkFrame(frame)
        database_name_frame.grid(row=2, column=0, columnspan=2)

        database_name_label = CTkLabel(database_name_frame, text="Nazwa bazy danych:")
        database_name_label.grid(row=0, column=0)

        database_name_entry = CTkEntry(database_name_frame, width=400, textvariable=self.__database_name__)
        database_name_entry.grid(row=1, column=0,padx=10, pady=10)

        default_data_button = CTkButton(frame, text="Wypełnij testowymi danymi", command=self.__set_default_data)
        default_data_button.grid(row=3, column=0, padx=10, pady=10, columnspan=2)

        test_connection = CTkButton(frame, text="Testuj połączenie", command=self.__test_connection__)
        test_connection.grid(row=4, column=0, padx=10, pady=10, columnspan=2)

        import_data_button = CTkButton(frame, text="Importuj wszystkie tabele", command=self.__import__all__tables__)
        import_data_button.grid(row=5, column=0, padx=10, pady=10, columnspan=2)

    def __create_style_form__(self, rootFrame: CTkFrame):
        frame = CTkFrame(rootFrame)
        frame.grid(row=1, column=1)

        is_align_content: BooleanVar = BooleanVar()
        is_add_colors: BooleanVar = BooleanVar()
        is_add_borders: BooleanVar = BooleanVar()


        # wyrównanie
        is_align_content_checkbox = CTkCheckBox(frame, text="Wyrównaj", variable=is_align_content, onvalue=True, offvalue=False)
        is_align_content_checkbox.grid(row=0, column=0, padx=10, pady=10, sticky="w")

        # dodanie kolorów do wierszy
        is_colors_checkbox = CTkCheckBox(frame, text="Dodaj kolory", variable=is_add_colors, onvalue=True, offvalue=False)
        is_colors_checkbox.grid(row=1, column=0, padx=10, pady=10, sticky="w")

        # dodanie borderów
        is_borders_checkbox = CTkCheckBox(frame, text="Dodaj krawędzie", variable=is_add_borders, onvalue=True, offvalue=False)
        is_borders_checkbox.grid(row=2, column=0, padx=10, pady=10, sticky="w")

        style_data_button = CTkButton(frame, text="Zastosuj stylowanie", command=lambda: self.__style_data__(is_align_content.get(), is_add_colors.get(), is_add_borders.get()))
        style_data_button.grid(row=3, column=0, padx=10, pady=10, columnspan=2)

    def __test_connection__(self):
        isValid = self.__validate_connect_form__()

        if isValid:
            try:
                connection = self.__connect_to_db__()
                messagebox.showinfo(title="Połączenie", message="Poprawnie nawiązano połączenie z bazą")
                connection.close()
            except:
                 messagebox.showerror("Błąd połączenia", message="Połączenie z bazą nie udane. Sprawdź poprawność danych.")
            

    def __validate_connect_form__(self):
            user_name = self.__user_name__.get()
            host = self.__host__.get()
            database_name = self.__database_name__.get()

            if not user_name:
                messagebox.showerror(title="Błąd walidacji", message="Nie podano nazwy użytkownika")
                return False
            
            if not host:
                 messagebox.showerror(title="Błąd walidacji", message="Nie podano hosta")
                 return False

            if not database_name:
                 messagebox.showerror(title="Błąd walidacji", message="Nie podano nazwy bazy danych")
                 return False
            
            return True

    def __connect_to_db__(self):
            user_name = self.__user_name__.get()
            user_password = self.__user_password__.get()
            host = self.__host__.get()
            database_name = self.__database_name__.get()
            
            db_connection = mysql.connector.connect(user=user_name,password=user_password, host=host, database=database_name)
            return db_connection
       

    def __import__all__tables__(self):
        try: 
            db_connection = self.__connect_to_db__()

            if db_connection:
                workbook = Workbook()
                self.__workbook__ = workbook
                cursor = db_connection.cursor()

                cursor.execute("SHOW TABLES;")

                tables = cursor.fetchall()

                for table in tables:
                    table_name = table[0]
                    cursor.execute(f"SELECT * FROM {table_name}")

                    table_data = cursor.fetchall()

                    ws = workbook.create_sheet(f'{table_name}-worksheet')

                    if ws and table_data: 
                        column_names = [column[0] for column in cursor.description]
                        ws.append(column_names)

                        for row in table_data:
                            ws.append(row)

            self.__on_safe__()
            db_connection.close()
        
        except:
            messagebox.showerror("Błąd połączenia", message="Połączenie z bazą nie udane. Sprawdź poprawność danych.")
    
    def __on_safe__(self):
        try:
            if self.__workbook__:
                filename = simpledialog.askstring("Filename", "Podaj nazwę pliku do jakiego zapisać")

                if filename:
                    self.__workbook__.save(f'./results/{filename}.xlsx')
                    self.__workbook_path__ = f'./results/{filename}.xlsx'
                    messagebox.showinfo(title="Zapisano", message="Poprawnie zapisano")
                else:
                    messagebox.showerror(title="Błąd zapisu", message="Nie wybrano nazwy pliku")
            else:
                raise Exception("No active workbook")
        except: 
            messagebox.showerror(title="Błąd zapisu", message="Bład zapisu")

    def __set_default_data(self):
        self.__user_name__.set(constant.DB_ROOT)
        self.__host__.set(constant.DB_HOST)
        self.__database_name__.set(constant.DB_NAME)

    def __align_content__(self):
         if self.__workbook__:
            ws = self.__workbook__.active
    
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column].width = adjusted_width

    def __add_colors__(self):
        if self.__workbook__:
            ws = self.__workbook__.active
            # Dodaj kolory dla pierwszego wiersza
            first_row_fill = PatternFill(start_color=constant.HEADER_COLOR, end_color=constant.HEADER_COLOR, fill_type='solid')
            for col in range(1, ws.max_column + 1):
                ws.cell(row=1, column=col).fill = first_row_fill

           # Dodaj różne kolory dla każdej kolumny (co druga kolumna inny kolor)
            for col in range(1, ws.max_column + 1):
                if col % 2 == 0:  # Jeśli kolumna jest parzysta
                    column_fill = PatternFill(start_color=constant.EVEN_COLUMN_COLOR, end_color=constant.EVEN_COLUMN_COLOR, fill_type='solid')  # Zielony kolor
                else:
                    column_fill = PatternFill(start_color=constant.ODD_COLUMN_COLOR, end_color=constant.ODD_COLUMN_COLOR, fill_type='solid')  # Żółty kolor
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=col).fill = column_fill
    
    def __add_borders(self):
        if self.__workbook__:
            ws = self.__workbook__.active

            thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin')) 

            for row in ws.iter_rows():
                for cell in row:
                    if cell.row == 1:  # Dla pierwszego wiersza (nagłówki)
                        cell.border = Border(left=Side(style='thick'), 
                                right=Side(style='thick'), 
                                bottom=Side(style='thick'))  # Grubszy border na dole
                    else:
                        cell.border = thin_border


    def __style_data__(self, is_align_content: bool, is_add_colors: bool,is_add_borders: bool):
        try:
   

            if  not self.__workbook__:
                file = filedialog.askopenfile(mode='r', initialdir="./", filetypes=[("xlsx", '*.xlsx'),('CSV', '*.csv')])
                if file: 
                    file_path = os.path.abspath(file.name)
                    self.__workbook__ = load_workbook(file_path)
                    self.__workbook_path__ = file_path
            
            if self.__workbook_path__:
                for sheet_name in self.__workbook__.sheetnames:
                    self.__workbook__.active = self.__workbook__[sheet_name]              
                    
                    if is_align_content:
                        self.__align_content__()

                    if is_add_colors:
                        self.__add_colors__()

                    if is_add_borders:
                        self.__add_borders()
       
                
                self.__workbook__.save(self.__workbook_path__)
                messagebox.showinfo(title="Poprawnie zapisano", message="Poprawnie zastosowano style")

            else:
                raise Exception("No file selected")
        except:
            messagebox.showerror(title="Błąd przy wyborze pliku", message="Nie wybrano pliku lub aktualnie wybrany plik jest już otwarty.")

    def open_window(self):
        super().open_window()
        self.__create_layout__()
